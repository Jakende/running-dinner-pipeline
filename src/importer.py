import csv
import json
import sqlite3
import time
import hashlib
import logging
import re
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime
from geopy.geocoders import Nominatim
from .models import Team, DietType

logger = logging.getLogger(__name__)

def _normalize_key(value: str) -> str:
    return " ".join(str(value).lower().strip().split())


class DinnerImporter:
    def __init__(self, input_file_path: str, db_path: str = 'dinner_data.db', mapping_path: str = 'field_mapping.json'):
        self.input_file_path = input_file_path
        self.json_file_path = input_file_path
        self.db_path = db_path
        self.geolocator = Nominatim(user_agent="freising_dinner_importer_v2")
        
        # Mapping constants (canonical fields -> supported LimeSurvey labels and question codes)
        self.KEYS = {
            'street': ["G01Q11[SQ001]", "Adresse euer Dinner-Location in Freising [Straße]"],
            'address_addition': ["G01Q11[SQ002]", "Adresse euer Dinner-Location in Freising [Adresszusatz]"],
            'house_number': ["G01Q11[SQ005]", "Adresse euer Dinner-Location in Freising [Hausnummer]"],
            'plz': ["G01Q11[SQ003]", "Adresse euer Dinner-Location in Freising [PLZ]"],
            'city': ["G01Q11[SQ004]", "Adresse euer Dinner-Location in Freising [Ort]"],
            'name1': ["G01Q08[SQ001]", "Namen [Name 1]"],
            'name2': ["G01Q08[SQ002]", "Namen [Name 2]"],
            'email1': ["G01Q04[SQ001]", "TUM/HSWT E-Mails [Email-1]", "TUM/HSWT E-Mail-Adressen [Email-1]"],
            'email2': ["G01Q04[SQ002]", "TUM/HSWT E-Mails [Email-2]", "TUM/HSWT E-Mail-Adressen [Email-2]"],
            'phone1': ["G01Q05[SQ001]", "Handynummer [Handynummer Teilnehmer:in 1]"],
            'phone2': ["G01Q05[SQ002]", "Handynummer [Handynummer Teilnehmer:in 2]"],
            'diet': ["G01Q02", "Eure Ernährungsweise ist ...? "],
            'hints': ["G01Q10", "Ein Wegweiser für eure Gäste"],
            'id': ["id", "Antwort ID"],
            'submitdate': ["submitdate", "Datum Abgeschickt"],
            'language': ["startlanguage", "Start-Sprache"]
        }
        
        self.ALLERGY_KEYS = {
            'gluten': ["G01Q07[SQ001]", "Allergene [Gluten]"],
            'lactose': ["G01Q07[SQ002]", "Allergene [Laktose]"],
            'peanuts': ["G01Q07[SQ003]", "Allergene [Erdnüsse]"],
            'soya': ["G01Q07[SQ004]", "Allergene [Soja]"],
            'nuts': ["G01Q07[SQ005]", "Allergene [Schalenfrüchte]"],
            'other': ["G01Q07[other]", "Allergene [Sonstiges]"]
        }
        self.custom_mapping = self._load_custom_mapping(mapping_path)

        self._init_db()

    def _load_custom_mapping(self, mapping_path: str) -> Dict[str, List[str]]:
        path = Path(mapping_path)
        if not path.exists():
            return {}
        with path.open('r', encoding='utf-8') as f:
            raw = json.load(f)
        mapping = {}
        for key, value in raw.items():
            mapping[key] = value if isinstance(value, list) else [value]
        logger.info(f"Loaded field mapping from {path}")
        return mapping

    def _init_db(self):
        """Initialize the SQLite database for caching geocoding results."""
        Path(self.db_path).parent.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS geocoding_cache (
                address_hash TEXT PRIMARY KEY,
                full_address TEXT,
                latitude REAL,
                longitude REAL,
                success INTEGER,
                updated_at TIMESTAMP
            )
        ''')
        conn.commit()
        conn.close()

    def _get_cached_location(self, address_hash: str) -> Optional[tuple]:
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT latitude, longitude, success FROM geocoding_cache WHERE address_hash = ?", (address_hash,))
        row = cursor.fetchone()
        conn.close()
        if row and row[2]: # if success
            return row[0], row[1]
        return None

    def _cache_location(self, address_hash: str, address: str, lat: Optional[float], lon: Optional[float], success: bool):
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO geocoding_cache (address_hash, full_address, latitude, longitude, success, updated_at)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (address_hash, address, lat, lon, 1 if success else 0, datetime.now().isoformat()))
        conn.commit()
        conn.close()

    def _geocode(self, address: str, address_hash: str) -> tuple:
        # Check cache
        cached = self._get_cached_location(address_hash)
        if cached:
            return cached[0], cached[1]

        # Geocode
        logger.info(f"Geocoding: {address}")
        try:
            location = self.geolocator.geocode(address)
            time.sleep(1.1)  # Respect rate limit
            if location:
                self._cache_location(address_hash, address, location.latitude, location.longitude, True)
                logger.info(f"  -> Found: {location.latitude}, {location.longitude}")
                return location.latitude, location.longitude
            else:
                self._cache_location(address_hash, address, None, None, False)
                logger.warning(f"  -> Not found: {address}")
                return None, None
        except Exception as e:
            logger.error(f"Geocoding error for {address}: {e}")
            return None, None

    def _load_responses(self) -> List[Dict]:
        path = Path(self.input_file_path)
        suffix = path.suffix.lower()
        logger.info(f"Loading data from {self.input_file_path}")

        if suffix == '.json':
            with path.open('r', encoding='utf-8') as f:
                data = json.load(f)
            return data.get('responses', []) if isinstance(data, dict) else data

        if suffix == '.csv':
            with path.open('r', encoding='utf-8-sig', newline='') as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample)
                except csv.Error:
                    dialect = csv.excel
                return list(csv.DictReader(f, dialect=dialect))

        if suffix == '.xlsx':
            try:
                from openpyxl import load_workbook
            except ImportError as exc:
                raise RuntimeError("XLSX import requires openpyxl. Install requirements.txt first.") from exc
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            headers = [str(v).strip() if v is not None else "" for v in next(rows)]
            responses = []
            for row in rows:
                responses.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
            return responses

        raise ValueError(f"Unsupported input format: {suffix}. Use .json, .csv, or .xlsx.")

    def _candidate_keys(self, logical_key: str, allergy: bool = False) -> List[str]:
        base = self.ALLERGY_KEYS if allergy else self.KEYS
        candidates = []
        candidates.extend(self.custom_mapping.get(logical_key, []))
        candidates.extend(base.get(logical_key, []))
        return [str(c) for c in candidates]

    def _get_value(self, row: Dict, logical_key: str, default: str = "", allergy: bool = False) -> str:
        candidates = self._candidate_keys(logical_key, allergy=allergy)

        for key in candidates:
            if key in row:
                return self._clean_value(row.get(key), default)

        candidate_norms = {_normalize_key(key) for key in candidates}
        for row_key, value in row.items():
            row_key_str = str(row_key)
            row_key_norm = _normalize_key(row_key_str)
            if row_key_norm in candidate_norms:
                return self._clean_value(value, default)

            # LimeSurvey exports often use "QCODE. Human readable label".
            if ". " in row_key_str:
                code, label = row_key_str.split(". ", 1)
                if _normalize_key(code) in candidate_norms or _normalize_key(label) in candidate_norms:
                    return self._clean_value(value, default)

        return default

    def _clean_value(self, value, default: str = "") -> str:
        if value is None:
            return default
        text = str(value).strip()
        if text in {"", "N/A"}:
            return default
        return text

    def _clean_email(self, value: str) -> str:
        return value.strip().lower() if value else ""

    def _clean_street(self, street: str) -> str:
        return (
            street.replace("strase", "straße")
            .replace("str.", "straße")
            .replace("Str.", "Straße")
            .replace("Haupstraße", "Hauptstraße")
            .replace("Dr.-von-Daller Straße", "Dr.-von-Daller-Straße")
            .strip()
        )

    def _clean_city(self, city: str) -> str:
        return city.replace(", Deutschland", "").replace("Deutschland", "").strip()

    def _clean_house_number(self, house: str, street: str) -> str:
        house = house.strip()
        if not house:
            return ""
        if house.lower() in {".", "-", "/", "_", "x", "xx", "n/a", "na"}:
            return ""
        if _normalize_key(house) == _normalize_key(street):
            return ""
        # Treat floor/WG notes as address hints, not house numbers.
        if re.search(r"(^|\b)(wg\s*\d*|eg|og|floor|geschoss|stock)\b", house, flags=re.IGNORECASE):
            return ""
        if not re.search(r"\d", house):
            return ""
        return house

    def _is_ignored_address_note(self, value: str, street: str = "") -> bool:
        text = value.strip()
        if not text:
            return True
        if text.lower() in {".", "-", "/", "_", "x", "xx", "n/a", "na"}:
            return True
        if street and _normalize_key(text) == _normalize_key(street):
            return True
        return False

    def _split_street_house(self, street: str) -> tuple[str, str]:
        street = street.strip()
        # Matches common German house numbers at the end, e.g. "Goethestraße 1",
        # "Main Street 12a", "Foo-Weg 7/1". Keeps the street name separate.
        match = re.match(r"^(?P<street>.*?\D)\s+(?P<house>\d+\s*[a-zA-Z]?(?:\s*[-/]\s*\d+\s*[a-zA-Z]?)?)$", street)
        if not match:
            return street, ""
        return match.group("street").strip(), re.sub(r"\s+", "", match.group("house").strip())

    def _build_address_parts(self, street: str, address_addition: str, house_field: str) -> tuple[str, str, str]:
        street = self._clean_street(street)
        street, house_from_street = self._split_street_house(street)
        addition_house = self._clean_house_number(address_addition, street)
        field_house = self._clean_house_number(house_field, street)

        house = house_from_street or addition_house or field_house
        notes = []

        if address_addition and address_addition != addition_house and not self._is_ignored_address_note(address_addition, street):
            notes.append(f"Adresszusatz: {address_addition}")
        elif (
            address_addition
            and house_from_street
            and _normalize_key(address_addition) != _normalize_key(house)
            and not self._is_ignored_address_note(address_addition, street)
        ):
            notes.append(f"Adresszusatz: {address_addition}")

        if house_field and house_field != field_house and not self._is_ignored_address_note(house_field, street):
            notes.append(house_field)
        elif (
            house_field
            and house != field_house
            and _normalize_key(house_field) != _normalize_key(house)
            and not self._is_ignored_address_note(house_field, street)
        ):
            notes.append(house_field)

        return street, house, "; ".join(notes)

    def _compose_address(self, street: str, house: str, plz: str, city: str) -> str:
        street_with_house = f"{street} {house}".strip()
        return f"{street_with_house}, {plz} {city}, Germany"

    def import_teams(self) -> List[Team]:
        responses = self._load_responses()
        teams = []

        for r in responses:
            response_id = self._get_value(r, 'id')
            if not self._get_value(r, 'submitdate'):
                continue # Skip incomplete
            
            # Extract Address
            street = self._get_value(r, 'street')
            address_addition = self._get_value(r, 'address_addition')
            house = self._get_value(r, 'house_number')
            plz = self._get_value(r, 'plz')
            city = self._get_value(r, 'city', "Freising")
            
            if not (street and plz):
                logger.warning(f"Skipping response {response_id}: Incomplete address")
                continue

            # Clean common typos and tolerate address forms where house numbers
            # were entered in the street or address-addition field.
            city = self._clean_city(city)
            street, house, address_notes = self._build_address_parts(street, address_addition, house)

            if not house:
                logger.warning(f"Response {response_id}: No house number detected; using street-level address")

            full_address = self._compose_address(street, house, plz, city)
            addr_hash = hashlib.md5(f"{street}{house}{plz}{city}".lower().encode()).hexdigest()
            
            lat, lon = self._geocode(full_address, addr_hash)
            if (lat is None or lon is None) and not house:
                approximate_address = self._compose_address(street, "", plz, city)
                approximate_hash = hashlib.md5(f"{street}{plz}{city}:approx".lower().encode()).hexdigest()
                logger.warning(
                    f"Precise geocoding failed for response {response_id}; "
                    f"trying approximate street-level address: {approximate_address}"
                )
                lat, lon = self._geocode(approximate_address, approximate_hash)
                if lat is not None and lon is not None:
                    full_address = approximate_address
            
            if lat is None or lon is None:
                logger.warning(f"Skipping response {response_id}: Geocoding failed")
                continue

            # diet
            diet_str = self._get_value(r, 'diet')
            diet = DietType.from_string(diet_str)

            # allergies
            allergies = {}
            for k in self.ALLERGY_KEYS:
                val = self._get_value(r, k, allergy=True)
                allergies[k] = val or None

            location_hints = "; ".join(
                part for part in [address_notes, self._get_value(r, 'hints')] if part
            )

            team = Team(
                id=int(response_id),
                name1=self._get_value(r, 'name1'),
                name2=self._get_value(r, 'name2'),
                email1=self._clean_email(self._get_value(r, 'email1')),
                email2=self._clean_email(self._get_value(r, 'email2')),
                phone1=self._get_value(r, 'phone1'),
                phone2=self._get_value(r, 'phone2'),
                address_street=f"{street} {house}",
                address_zip=plz,
                address_city=city,
                full_address=full_address,
                latitude=lat,
                longitude=lon,
                diet=diet,
                allergies=allergies,
                hints=location_hints,
                language=self._get_value(r, 'language', "de")
            )
            teams.append(team)

        logger.info(f"Imported {len(teams)} valid teams")
        return teams
